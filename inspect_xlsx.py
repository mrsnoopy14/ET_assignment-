from __future__ import annotations

from openpyxl import load_workbook


def main() -> int:
    wb = load_workbook("3. Heat_Sink_Design_Ref.xlsx", data_only=False)
    print("sheets:", wb.sheetnames)

    keywords = [
        "Junction",
        "Total Heat Sink",
        "Resitance",  # misspelling present in PDF
        "Resistance",
        "Rt",
        "Rhs",
    ]

    for name in wb.sheetnames:
        ws = wb[name]
        hits: list[tuple[str, str]] = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(k in v for k in keywords):
                    hits.append((cell.coordinate, v))

        if hits:
            print("\nSheet", name)
            for coord, v in hits[:50]:
                print(f"  {coord}: {v}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
