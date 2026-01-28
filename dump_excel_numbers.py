from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    wb = load_workbook("3. Heat_Sink_Design_Ref.xlsx", data_only=True)

    out_dir = Path("out")
    out_dir.mkdir(exist_ok=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    rows.append(f"{sheet}!{cell.coordinate}\t{cell.value}")

        (out_dir / f"excel_numbers_{sheet}.tsv").write_text("\n".join(rows), encoding="utf-8")
        print(sheet, "numbers:", len(rows))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
