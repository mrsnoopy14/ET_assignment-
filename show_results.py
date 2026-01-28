from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    root = Path(__file__).resolve().parent
    sys.path.insert(0, str(root / "src"))

    from thermal_model import ThermalInputs, results_as_dict, solve_thermal_model  # noqa: E402

    wb = load_workbook(str(root / "3. Heat_Sink_Design_Ref.xlsx"), data_only=True)
    ws = wb["Sheet1"]

    excel = {
        "r_total_c_w": ws["I37"].value,
        "t_junction_c": ws["I38"].value,
        "r_conv_c_w": ws["I30"].value,
        "r_cond_c_w": ws["I29"].value,
        "r_tim_c_w": ws["I26"].value,
        "r_jc_c_w": ws["I22"].value,
        "die_area_m2": ws["I25"].value,
    }

    inputs = ThermalInputs()
    results = solve_thermal_model(inputs)

    report = {
        "inputs": inputs.__dict__,
        "python_results": results_as_dict(results),
        "excel_reference_cells": {
            "Sheet1!I37 (R_total)": excel["r_total_c_w"],
            "Sheet1!I38 (T_junction)": excel["t_junction_c"],
            "Sheet1!I30 (R_conv)": excel["r_conv_c_w"],
            "Sheet1!I29 (R_cond)": excel["r_cond_c_w"],
            "Sheet1!I26 (R_tim)": excel["r_tim_c_w"],
            "Sheet1!I22 (R_jc)": excel["r_jc_c_w"],
            "Sheet1!I25 (A_die)": excel["die_area_m2"],
        },
        "key_results": {
            "R_total_C_per_W": results.r_total_c_w,
            "T_junction_C": results.t_junction_c,
            "R_hs_C_per_W": results.r_hs_c_w,
        },
    }

    out_dir = root / "out"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "show_results.json"
    out_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
