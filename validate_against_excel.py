from __future__ import annotations

import argparse
import math
import sys
from dataclasses import asdict
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


def _add_src_to_path() -> None:
    root = Path(__file__).resolve().parent
    src = root / "src"
    sys.path.insert(0, str(src))


def _is_number(v: object) -> bool:
    return isinstance(v, (int, float)) and not (isinstance(v, float) and (math.isnan(v) or math.isinf(v)))


def _neighbors(ws, cell):
    r, c = cell.row, cell.column
    coords = [
        (r, c + 1),
        (r, c + 2),
        (r + 1, c),
        (r + 2, c),
        (r + 1, c + 1),
    ]
    for rr, cc in coords:
        try:
            yield ws.cell(row=rr, column=cc)
        except Exception:
            continue


def _find_labeled_numeric(
    wb,
    label_contains: Iterable[str],
    *,
    expected_min: float,
    expected_max: float,
) -> list[dict[str, object]]:
    needles = [s.lower() for s in label_contains]
    matches: list[dict[str, object]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                text = " ".join(v.split()).lower()
                if not all(n in text for n in needles):
                    continue

                for nb in _neighbors(ws, cell):
                    if _is_number(nb.value):
                        num = float(nb.value)
                        in_range = expected_min <= num <= expected_max
                        matches.append(
                            {
                                "sheet": sheet_name,
                                "label_cell": cell.coordinate,
                                "label": v,
                                "value_cell": nb.coordinate,
                                "value": num,
                                "in_range": in_range,
                            }
                        )

    # Prefer in-range matches, then earlier ones
    matches.sort(key=lambda m: (not bool(m["in_range"]), str(m["sheet"]), str(m["value_cell"])))
    return matches


def _pick_best(matches: list[dict[str, object]]) -> Optional[dict[str, object]]:
    return matches[0] if matches else None


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate the thermal Python model against the provided Excel reference.")
    parser.add_argument(
        "--xlsx",
        default="3. Heat_Sink_Design_Ref.xlsx",
        help="Path to the reference workbook (default: %(default)s)",
    )
    parser.add_argument("--tol-r", type=float, default=0.02, help="Abs tolerance for R_hs (degC/W)")
    parser.add_argument("--tol-t", type=float, default=1.0, help="Abs tolerance for T_junction (degC)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: workbook not found: {xlsx_path}")
        return 2

    wb = load_workbook(str(xlsx_path), data_only=True)

    # The provided reference workbook stores the final outputs on Sheet1 at:
    # - Total resistance (reported as "Total Heat Sink Resitance" in the PDF): I37
    # - Junction temperature: I38
    # These are cached numeric values, so reading by coordinate is the most reliable.
    sheet_name = "Sheet1"
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Expected sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
        return 2
    ws = wb[sheet_name]

    excel_r_total = ws["I37"].value
    excel_t_junction = ws["I38"].value
    excel_r_conv = ws["I30"].value
    excel_r_cond = ws["I29"].value
    excel_r_tim = ws["I26"].value

    _add_src_to_path()
    from thermal_model import ThermalInputs, solve_thermal_model  # noqa: E402

    inputs = ThermalInputs()
    results = solve_thermal_model(inputs)

    print("Python model (defaults aligned to Excel reference):")
    print("  r_total_c_w   =", results.r_total_c_w)
    print("  t_junction_c  =", results.t_junction_c)
    print("  r_conv_c_w    =", results.r_conv_c_w)
    print("  r_cond_c_w    =", results.r_cond_c_w)
    print("  r_tim_c_w     =", results.r_tim_c_w)

    ok = True

    def require_num(name: str, v: object) -> float:
        if v is None:
            raise ValueError(f"Excel cell for {name} is empty (missing cached value).")
        if not _is_number(v):
            raise ValueError(f"Excel cell for {name} is not numeric: {v!r}")
        return float(v)

    try:
        excel_r_total_f = require_num("R_total (Sheet1!I37)", excel_r_total)
        excel_t_junction_f = require_num("T_junction (Sheet1!I38)", excel_t_junction)
        excel_r_conv_f = require_num("R_conv (Sheet1!I30)", excel_r_conv)
        excel_r_cond_f = require_num("R_cond (Sheet1!I29)", excel_r_cond)
        excel_r_tim_f = require_num("R_tim (Sheet1!I26)", excel_r_tim)
    except ValueError as e:
        print("\nFAIL:", e)
        print("Tip: open the workbook in Excel, hit Recalculate, save, then re-run.")
        return 1

    def check(name: str, py: float, xl: float, tol: float) -> None:
        nonlocal ok
        diff = abs(py - xl)
        print(f"\n{name}:")
        print("  excel =", xl)
        print("  python=", py)
        print("  abs diff=", diff)
        if diff > tol:
            ok = False
            print(f"  FAIL: exceeds tol ({tol})")
        else:
            print("  PASS")

    check("R_total (Sheet1!I37)", results.r_total_c_w, excel_r_total_f, args.tol_r)
    check("T_junction (Sheet1!I38)", results.t_junction_c, excel_t_junction_f, args.tol_t)
    # These should be very tight; keep them under tol_r for simplicity.
    check("R_conv (Sheet1!I30)", results.r_conv_c_w, excel_r_conv_f, args.tol_r)
    check("R_cond (Sheet1!I29)", results.r_cond_c_w, excel_r_cond_f, args.tol_r)
    check("R_tim (Sheet1!I26)", results.r_tim_c_w, excel_r_tim_f, args.tol_r)

    if not ok:
        print("\nIf this fails but your equations match the PDF, the most common causes are:")
        print("- Excel cached formula results are missing (open in Excel, recalc, save)")
        print("- Different convection area definition / fin efficiency assumptions in the sheet")
        print("- Unit mismatch (mm vs m) in one of the terms")
        return 1

    print("\nOK: Python model matches Excel within tolerances.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
